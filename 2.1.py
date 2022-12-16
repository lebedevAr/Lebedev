import csv
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Border, Side


class Vacancy:
    to_rub_dict = {"AZN": 35.68, "BYR": 23.91, "EUR": 59.90, "GEL": 21.74, "KGS": 0.76, "KZT": 0.13, "RUR": 1,
                   "UAH": 1.64, "USD": 60.66, "UZS": 0.0055}

    def __init__(self, vacancy):
        self.name = vacancy['name']
        self.salary_from = int(float(vacancy['salary_from']))
        self.salary_to = int(float(vacancy['salary_to']))
        self.salary_currency = vacancy['salary_currency']
        self.salary_average = self.to_rub_dict[self.salary_currency] * (self.salary_from + self.salary_to) / 2
        self.area_name = vacancy['area_name']
        self.year = int(vacancy['published_at'][:4])


class DataSet:
    def __init__(self, file_name, vacancy_name):
        self.file_name = file_name
        self.vacancy_name = vacancy_name

    def csv_reader(self):
        with open(self.file_name, mode='r', encoding='utf-8-sig') as file:
            reader = csv.reader(file)
            header = next(reader)
            header_length = len(header)
            for row in reader:
                if '' not in row and len(row) == header_length:
                    yield dict(zip(header, row))

    @staticmethod
    def avg(dictionary):
        new_dict = {}
        for key, values in dictionary.items():
            new_dict[key] = int(sum(values) / len(values))
        return new_dict

    @staticmethod
    def increment(dictionary, key, amount):
        if key in dictionary:
            dictionary[key] += amount
        else:
            dictionary[key] = amount

    def get_stats(self):
        salary = {}
        salary_of_vacancy = {}
        salary_city = {}
        vacancies_counter = 0

        new_csv = self.csv_reader()

        for vacancy_dictionary in new_csv:
            vacancy = Vacancy(vacancy_dictionary)
            self.increment(salary, vacancy.year, [vacancy.salary_average])
            if vacancy.name.find(self.vacancy_name) != -1:
                self.increment(salary_of_vacancy, vacancy.year, [vacancy.salary_average])
            self.increment(salary_city, vacancy.area_name, [vacancy.salary_average])
            vacancies_counter += 1

        vacancies_num = dict([(key, len(value)) for key, value in salary.items()])
        number_by_name = dict([(key, len(value)) for key, value in salary_of_vacancy.items()])

        if not salary_of_vacancy:
            salary_of_vacancy = dict([(key, [0]) for key, value in salary.items()])
            number_by_name = dict([(key, 0) for key, value in vacancies_num.items()])

        stats = self.avg(salary)
        stats2 = self.avg(salary_of_vacancy)
        stats3 = self.avg(salary_city)

        stats4 = {}
        for year, salaries in salary_city.items():
            stats4[year] = round(len(salaries) / vacancies_counter, 4)
        stats4 = list(filter(lambda a: a[-1] >= 0.01, [(key, value) for key, value in stats4.items()]))
        stats4.sort(key=lambda a: a[-1], reverse=True)
        stats5 = stats4.copy()
        stats4 = dict(stats4)
        stats3 = list(filter(lambda a: a[0] in list(stats4.keys()), [(key, value) for key, value in stats3.items()]))
        stats3.sort(key=lambda a: a[-1], reverse=True)
        stats3 = dict(stats3[:10])
        stats5 = dict(stats5[:10])

        return stats, vacancies_num, stats2, number_by_name, stats3, stats5

    @staticmethod
    def print_statistic(stats1, stats2, stats3, stats4, stats5, stats6):
        print('Динамика уровня зарплат по годам: ' + str(stats1))
        print('Динамика количества вакансий по годам: ' + str(stats2))
        print('Динамика уровня зарплат по годам для выбранной профессии: ' + str(stats3))
        print('Динамика количества вакансий по годам для выбранной профессии: ' + str(stats4))
        print('Уровень зарплат по городам (в порядке убывания): ' + str(stats5))
        print('Доля вакансий по городам (в порядке убывания): ' + str(stats6))


class InputConnect:
    def __init__(self):
        self.file_name = input('Введите название файла: ')
        self.vacancy_name = input('Введите название профессии: ')

        dataset = DataSet(self.file_name, self.vacancy_name)
        stats1, stats2, stats3, stats4, stats5, stats6 = dataset.get_stats()
        dataset.print_statistic(stats1, stats2, stats3, stats4, stats5, stats6)

        new_report = Report(self.vacancy_name, stats1, stats2, stats3, stats4, stats5, stats6)
        ws1 = new_report.get_first_sheet()
        ws2 = new_report.get_second_sheet()
        new_report.generate_excel(ws1, ws2)


class Report:
    def __init__(self, vacancy_name, stats1, stats2, stats3, stats4, stats5, stats6):
        self.wb = Workbook()
        self.vacancy_name = vacancy_name
        self.stats1 = stats1
        self.stats2 = stats2
        self.stats3 = stats3
        self.stats4 = stats4
        self.stats5 = stats5
        self.stats6 = stats6

    def get_first_sheet(self):
        ws1 = self.wb.active
        ws1.title = 'Статистика по годам'
        ws1.append(['Год', 'Средняя зарплата', 'Средняя зарплата - ' + self.vacancy_name, 'Количество вакансий',
                    'Количество вакансий - ' + self.vacancy_name])
        for year in self.stats1.keys():
            ws1.append([year, self.stats1[year], self.stats3[year], self.stats2[year], self.stats4[year]])

        data = [['Год ', 'Средняя зарплата ', ' Средняя зарплата - ' + self.vacancy_name, ' Количество вакансий',
                 ' Количество вакансий - ' + self.vacancy_name]]
        column_widths = []
        for row in data:
            for i, cell in enumerate(row):
                if len(column_widths) > i:
                    if len(cell) > column_widths[i]:
                        column_widths[i] = len(cell)
                else:
                    column_widths += [len(cell)]

        for i, column_width in enumerate(column_widths, 1):  # ,1 to start at 1
            ws1.column_dimensions[get_column_letter(i)].width = column_width + 2
        return ws1

    def get_second_sheet(self):
        new_data = [['Город', 'Уровень зарплат', '', 'Город', 'Доля вакансий']]

        for (city1, value1), (city2, value2) in zip(self.stats5.items(), self.stats6.items()):
            new_data.append([city1, value1, '', city2, value2])
        ws2 = self.wb.create_sheet('Статистика по городам')
        for row in new_data:
            ws2.append(row)

        column_widths = []
        for row in new_data:
            for i, cell in enumerate(row):
                cell = str(cell)
                if len(column_widths) > i:
                    if len(cell) > column_widths[i]:
                        column_widths[i] = len(cell)
                else:
                    column_widths += [len(cell)]

        for i, column_width in enumerate(column_widths, 1):
            ws2.column_dimensions[get_column_letter(i)].width = column_width + 2

        return ws2

    def generate_excel(self, ws1, ws2):
        font_bold = Font(bold=True)
        for colum in 'ABCDE':
            ws1[colum + '1'].font = font_bold
            ws2[colum + '1'].font = font_bold

        for index, _ in enumerate(self.stats5):
            ws2['E' + str(index + 2)].number_format = '0.00%'

        thin = Side(border_style='thin', color='00000000')

        self.stats1[1] = 1
        for row, _ in enumerate(self.stats1):
            for colum in 'ABCDE':
                ws1[colum + str(row + 1)].border = Border(left=thin, bottom=thin, right=thin, top=thin)

        for row in range(len(self.stats6.keys()) + 1):
            for colum in 'ABDE':
                ws2[colum + str(row + 1)].border = Border(left=thin, bottom=thin, right=thin, top=thin)

        self.wb.save('report.xlsx')


if __name__ == '__main__':
    InputConnect()
