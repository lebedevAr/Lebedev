import csv
import pathlib
import pdfkit

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Border, Side
from jinja2 import Environment, PackageLoader, select_autoescape, FileSystemLoader

import matplotlib.pyplot as plt
import numpy as np


class Vacancy:
    """Класс для представления вакансии.

    Attributes:
        name (str): Название вакансии
        salary_from (int): Нижняя гранциа вилки оклада
        salary_to (int): Верхняя гранциа вилки оклада
        salary_currency (str): Значение валюты оклада
        salary_average (int): Среднее значение зарплаты
        area_name (str): Город, указанный в вакансии
        year (int): Год публикации вакансии.
    """
    to_rub_dict = {"AZN": 35.68, "BYR": 23.91, "EUR": 59.90, "GEL": 21.74, "KGS": 0.76, "KZT": 0.13, "RUR": 1,
                   "UAH": 1.64, "USD": 60.66, "UZS": 0.0055}

    def __init__(self, vacancy):
        """Инициализирует объект Vacancy, выполняет конвертации для целочисленных полей.

        Args:
            vacancy (dict): Словарь содержащий поля вакансии, необходиме для инициализации полей объекта Vacancy.

        >>> type(Vacancy({'name': 'Программист', 'salary_from': '20000', 'salary_to': '30000', 'salary_currency': 'RUR', 'area_name': 'Екатеринбург', 'published_at': '2020-11-03T17:11:35+0300'})).__name__
		'Vacancy'
		>>> Vacancy({'name': 'Программист', 'salary_from': '20000', 'salary_to': '30000', 'salary_currency': 'RUR', 'area_name': 'Екатеринбург', 'published_at': '2020-11-03T17:11:35+0300'}).name
		'Программист'
		>>> Vacancy({'name': 'Программист', 'salary_from': '20000', 'salary_to': '30000', 'salary_currency': 'RUR', 'area_name': 'Екатеринбург', 'published_at': '2020-11-03T17:11:35+0300'}).area_name
		'Екатеринбург'
		>>> Vacancy({'name': 'Программист', 'salary_from': '20000', 'salary_to': '30000', 'salary_currency': 'RUR', 'area_name': 'Екатеринбург', 'published_at': '2020-11-03T17:11:35+0300'}).year
		'2020'
		"""

        self.name = vacancy['name']
        self.salary_from = int(float(vacancy['salary_from']))
        self.salary_to = int(float(vacancy['salary_to']))
        self.salary_currency = vacancy['salary_currency']
        self.salary_average = self.to_rub_dict[self.salary_currency] * (self.salary_from + self.salary_to) / 2
        self.area_name = vacancy['area_name']
        self.year = int(vacancy['published_at'][:4])


class DataSet:
    """Класс для обработки и хранения данных из csv файла.

    Attributes:
        file_name (str): название csv файла, расположенного в папке проекта
        vacancy_name (str): Название вакансии, по которой требуется.
    """

    def __init__(self, file_name, vacancy_name):
        """Инициализирует объект Vacancy, выполняет конвертации для целочисленных полей.

        Args:
            file_name (str): название csv файла, расположенного в папке проекта
            vacancy_name (str): Название вакансии, по которой требуется.

        >>> DataSet('vacancies_by_year.csv', 'Программист').file_name
		'vacancies_by_year.csv'
		>>> DataSet('vacancies.csv', 'Программист').file_name
		'vacancies.csv'
		>>> DataSet('vacancies.csv', 'Аналитик').vacancy_name
		'Аналитик'
		>>> DataSet('vacancies.csv', 'Программист').vacancy_name
		'Программист'
		"""

        self.file_name = file_name
        self.vacancy_name = vacancy_name

    def csv_reader(self):
        """Читает csv файл и выделяет названия полей.

        Returns:
            dict: dict (key - названия полей, value - значения полей).
        """
        with open(self.file_name, mode='r', encoding='utf-8-sig') as file:
            reader = csv.reader(file)
            header = next(reader)
            header_length = len(header)
            for row in reader:
                if '' not in row and len(row) == header_length:
                    yield dict(zip(header, row))

    @staticmethod
    def avg(dictionary):
        """Создает словарь со средним значение по годам.

        Args:
            dictionary (dict): Словарь (key - год, value - list значений).

        Returns:
            1) (dict): dict(key - год, value - среднее значение за год).
        """
        new_dict = {}
        for key, values in dictionary.items():
            new_dict[key] = int(sum(values) / len(values))
        return new_dict

    @staticmethod
    def increment(dictionary, key, amount):
        """Служит для сбора статистики по ключу.

        Args:
            dictionary (dict): словарь, в который сохраняется статистика
            key (int): Ключ нужного элемента
            amount (list[int]): Количество элементов.
        """
        if key in dictionary:
            dictionary[key] += amount
        else:
            dictionary[key] = amount

    def get_stats(self):
        """Собирает полную статистику по csv файлу, необходимую для дальнейшей работы программы.

        Returns:
            1) (dict): key - год, value - средняя з/п по всем вакансиям
		    2) (dict): key - год, value - количество вакансий
		    3) (dict): key - город, value - средняя з/п в Городе
		    4) (dict): key - год, value - количество вакансий по выбранной профессии
		    5) (dict): key - Город, value - средняя з/п
		    6) (dict): key - Город, value - доля по выбранной вакансии от общего количества
		"""

        salary = {}
        salary_of_vacancy = {}
        salary_city = {}
        vacancies_counter = 0

        new_csv = self.csv_reader()

        for vacancy_dictionary in new_csv:
            vacancy = Vacancy(vacancy_dictionary)
            self.increment(salary, vacancy.year, [vacancy.salary_average])
            if vacancy.name.find(self.vacancy_name) != -1:
                self.increment(salary_of_vacancy, vacancy.year, [vacancy.salary_average])
            self.increment(salary_city, vacancy.area_name, [vacancy.salary_average])
            vacancies_counter += 1

        vacancies_num = dict([(key, len(value)) for key, value in salary.items()])
        number_by_name = dict([(key, len(value)) for key, value in salary_of_vacancy.items()])

        if not salary_of_vacancy:
            salary_of_vacancy = dict([(key, [0]) for key, value in salary.items()])
            number_by_name = dict([(key, 0) for key, value in vacancies_num.items()])

        stats = self.avg(salary)
        stats2 = self.avg(salary_of_vacancy)
        stats3 = self.avg(salary_city)

        stats4 = {}
        for year, salaries in salary_city.items():
            stats4[year] = round(len(salaries) / vacancies_counter, 4)
        stats4 = list(filter(lambda a: a[-1] >= 0.01, [(key, value) for key, value in stats4.items()]))
        stats4.sort(key=lambda a: a[-1], reverse=True)
        stats5 = stats4.copy()
        stats4 = dict(stats4)
        stats3 = list(filter(lambda a: a[0] in list(stats4.keys()), [(key, value) for key, value in stats3.items()]))
        stats3.sort(key=lambda a: a[-1], reverse=True)
        stats3 = dict(stats3[:10])
        stats5 = dict(stats5[:10])

        return stats, vacancies_num, stats2, number_by_name, stats3, stats5

    @staticmethod
    def print_statistic(stats1, stats2, stats3, stats4, stats5, stats6):
        """Печатает статистику в консоль

        Args:
            stats1 (dict): dict(key - год, value - средняя з/п по всем вакансиям)
            stats2 (dict): dict(key - год, value - количество вакансий)
            stats3 (dict): dict(key - город, value - средняя з/п в Городе)
            stats4 (dict): dict(key - год, value - количество вакансий по выбранной профессии)
            stats5 (dict): dict(key - Город, value - средняя з/п)
            stats6 (dict): dict(key - Город, value - доля по выбранной вакансии от общего количества).
        """
        print('Динамика уровня зарплат по годам: ' + str(stats1))
        print('Динамика количества вакансий по годам: ' + str(stats2))
        print('Динамика уровня зарплат по годам для выбранной профессии: ' + str(stats3))
        print('Динамика количества вакансий по годам для выбранной профессии: ' + str(stats4))
        print('Уровень зарплат по городам (в порядке убывания): ' + str(stats5))
        print('Доля вакансий по городам (в порядке убывания): ' + str(stats6))


class InputConnect:
    """Класс для сохранения отчёта.

    Attributes:
        file_name (str): название csv файла, расположенного в папке проекта
        vacancy_name (str): Название вакансии, по которой требуется.
        user_select (str): Ключевое слово, определяющее в каком виде сохранить отчёт
    """

    def __init__(self):
        """Инициализирует объект InputConnect, выполняет сохранение отчёта в зависимости от выбранного типа."""

        self.file_name = input('Введите название файла: ')
        self.vacancy_name = input('Введите название профессии: ')
        self.user_select = input('Вакансии или статистика?: ')
        dataset = DataSet(self.file_name, self.vacancy_name)
        stats1, stats2, stats3, stats4, stats5, stats6 = dataset.get_stats()
        # dataset.print_statistic(stats1, stats2, stats3, stats4, stats5, stats6)
        new_graphic = Report(self.vacancy_name, stats1, stats2, stats3, stats4, stats5, stats6)

        if self.user_select.lower() == 'вакансии':
            ws1 = new_graphic.get_first_sheet()
            ws2 = new_graphic.get_second_sheet()
            new_graphic.generate_excel(ws1, ws2)
        elif self.user_select.lower() == 'статистика':
            new_graphic.generate_image()

        # new_graphic.generate_pdf()


class Report:
    """Класс для формирования отчёта.

        Attributes:
            wb (Workbook): название csv файла, расположенного в папке проекта
            vacancy_name (str): Название вакансии, по которой требуется
            stats1 (dict): key - год, value - средняя з/п по всем вакансиям
            stats2 (dict): key - год, value - количество вакансий
            stats3 (dict): key - город, value - средняя з/п в Городе
            stats4 (dict): key - год, value - количество вакансий по выбранной профессии
            stats5 (dict): key - Город, value - средняя з/п
            stats6 (dict): key - Город, value - доля по выбранной вакансии от общего количества.
        """

    def __init__(self, vacancy_name, stats1, stats2, stats3, stats4, stats5, stats6):
        """Инициализирует объект InputConnect.

        >>> Report({'2020': 30000, '2019': 35000}, {}, {}, {}, {}, {}).salary['2019']
		35000
		>>> Report({}, {'2020': 1000, '2019': 700}, {}, {}, {}, {}).amount['2020']
		1000
		>>> Report({}, {}, {}, {}, {}, {'Москва': 0.23, 'Екатеринбург': 0.11}).stats6['Москва']
		0.23
		"""

        self.wb = Workbook()
        self.vacancy_name = vacancy_name
        self.stats1 = stats1
        self.stats2 = stats2
        self.stats3 = stats3
        self.stats4 = stats4
        self.stats5 = stats5
        self.stats6 = stats6

    def generate_image(self):
        """Создаёт и сохраняет изображение с 4 графиками статистики."""
        fig, ax = plt.subplots(2, 2)
        x_axis = np.arange(len(self.stats1.keys()))

        data = self.stats5
        courses = [label.replace(' ', '\n').replace('-', '-\n') for label in list(data.keys())]
        values = list(data.values())

        def get_columns():
            """Распределяет названия столбцов

            Returns:
            labels (list[str]): Названия городов
                sizes (list[int]): Доли профессий на город
            """
            new_dic = {'Другие': 1 - sum((list(self.stats6.values())))}
            new_dic.update(self.stats6)
            stats6 = new_dic
            labels = list(stats6.keys())
            sizes = list(stats6.values())
            return labels, sizes

        def get_graph(i, stats1, stats2, name, text):
            """Создаёт график под определенный номер

            Args:
                i (int): номер графика
                stats1 (dict): данные для графика
                stats2 (dict): данные для графика
                name (str): Название профессии
                text (list[str]): Текст для графика
            """

            w = 0.4
            ax[0, i].bar(x_axis - w / 2, stats1.values(), width=w, label=text[0])
            ax[0, i].bar(x_axis + w / 2, stats2.values(), width=w, label=text[1] + str(name))
            ax[0, i].set_xticks(x_axis, stats1.keys())
            ax[0, i].set_xticklabels(stats1.keys(), rotation='vertical', va='top', ha='center')
            ax[0, i].set_title(text[2])
            ax[0, i].grid(True, axis='y')
            ax[0, i].tick_params(axis='both', labelsize=8)
            ax[0, i].legend(fontsize=8)

        get_graph(0, self.stats1, self.stats3, self.vacancy_name, ["средняя з/п", 'з/п ', 'Уровень зарплат по годам'])
        get_graph(1, self.stats2, self.stats4, self.vacancy_name, ['Количество вакансий', 'Количество вакансий\n',
                                                                   'Количество вакансий по годам'])

        ax[1, 0].set_title("Уровень зарплат по городам")
        ax[1, 0].invert_yaxis()
        ax[1, 0].tick_params(axis='both', labelsize=8)
        ax[1, 0].set_yticklabels(courses, fontsize=6, va='center', ha='right')
        ax[1, 0].barh(courses, values)
        ax[1, 0].grid(True, axis='x')

        labels, sizes = get_columns()
        ax[1, 1].pie(sizes, labels=labels, textprops={'fontsize': 6})
        ax[1, 1].axis('scaled')
        ax[1, 1].set_title("Доля вакансий по городам")

        plt.tight_layout()
        plt.savefig('graph.png', dpi=300)

    def get_first_sheet(self):
        """Создаёт объект Excel worksheet

        Returns:
            ws1 : Первый лист объекта
        """
        ws1 = self.wb.active
        ws1.title = 'Статистика по годам'
        ws1.append(['Год', 'Средняя зарплата', 'Средняя зарплата - ' + self.vacancy_name, 'Количество вакансий',
                    'Количество вакансий - ' + self.vacancy_name])
        for year in self.stats1.keys():
            ws1.append([year, self.stats1[year], self.stats3[year], self.stats2[year], self.stats4[year]])

        data = [['Год ', 'Средняя зарплата ', ' Средняя зарплата - ' + self.vacancy_name, ' Количество вакансий',
                 ' Количество вакансий - ' + self.vacancy_name]]
        column_widths = []
        for row in data:
            for i, cell in enumerate(row):
                if len(column_widths) > i:
                    if len(cell) > column_widths[i]:
                        column_widths[i] = len(cell)
                else:
                    column_widths += [len(cell)]

        for i, column_width in enumerate(column_widths, 1):  # ,1 to start at 1
            ws1.column_dimensions[get_column_letter(i)].width = column_width + 2
        return ws1

    def get_second_sheet(self):
        """Создаёт объект Excel worksheet

        Returns:
            ws2 : Второй лист объекта
        """
        new_data = [['Город', 'Уровень зарплат', '', 'Город', 'Доля вакансий']]

        for (city1, value1), (city2, value2) in zip(self.stats5.items(), self.stats6.items()):
            new_data.append([city1, value1, '', city2, value2])
        ws2 = self.wb.create_sheet('Статистика по городам')
        for row in new_data:
            ws2.append(row)

        column_widths = []
        for row in new_data:
            for i, cell in enumerate(row):
                cell = str(cell)
                if len(column_widths) > i:
                    if len(cell) > column_widths[i]:
                        column_widths[i] = len(cell)
                else:
                    column_widths += [len(cell)]

        for i, column_width in enumerate(column_widths, 1):
            ws2.column_dimensions[get_column_letter(i)].width = column_width + 2

        return ws2

    def generate_excel(self, ws1, ws2):
        """Склеивает 2 листа в один файл Excel и сохраняет со всеми настройками

        Args:
            ws1 : Первый лист Excel
            ws2 : Второй лист Excel
        """

        font_bold = Font(bold=True)
        for colum in 'ABCDE':
            ws1[colum + '1'].font = font_bold
            ws2[colum + '1'].font = font_bold

        for index, _ in enumerate(self.stats5):
            ws2['E' + str(index + 2)].number_format = '0.00%'

        thin = Side(border_style='thin', color='00000000')

        self.stats1[1] = 1
        for row, _ in enumerate(self.stats1):
            for colum in 'ABCDE':
                ws1[colum + str(row + 1)].border = Border(left=thin, bottom=thin, right=thin, top=thin)

        for row in range(len(self.stats6.keys()) + 1):
            for colum in 'ABDE':
                ws2[colum + str(row + 1)].border = Border(left=thin, bottom=thin, right=thin, top=thin)

        self.wb.save('report.xlsx')

    def generate_pdf(self):
        """Совмещает в себе статистику на графиках и в таблицах, сохраняет файл в формате pdf"""
        env = Environment(loader=FileSystemLoader('.'))
        template = env.get_template("pdf_template.html")
        stats = []
        for year in self.stats2.keys():
            stats.append([year, self.stats1[year], self.stats2[year], self.stats3[year], self.stats4[year]])

        for key in self.stats6:
            self.stats6[key] = round(self.stats6[key] * 100, 2)

        pdf_template = template.render({'name': self.vacancy_name,
                                        'path': '{0}/{1}'.format(pathlib.Path(__file__).parent.resolve(), 'graph.png'),
                                        'stats': stats, 'stats5': self.stats5, 'stats6': self.stats6})

        config = pdfkit.configuration(wkhtmltopdf=r'C:\Program Files\wkhtmltopdf\bin\wkhtmltopdf.exe')
        pdfkit.from_string(pdf_template, 'report.pdf', configuration=config, options={"enable-local-file-access": ""})


if __name__ == '__main__':
    InputConnect()
